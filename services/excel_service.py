"""
services.excel_service
======================
All Excel I/O for the application — **no Flask dependencies**.

Responsibilities
----------------
- **Import**   Parse an uploaded ``.xlsx`` / ``.xls`` file and
               upsert Invoice records into the database.
- **Export**   Build an ``openpyxl`` ``Workbook`` for the GST monthly
               report (caller turns it into a Flask ``send_file`` response).
- **Template** Build the blank invoice-upload template workbook.

Required pip package: ``openpyxl``  (``pip install openpyxl``)
"""
from __future__ import annotations

import logging
from calendar import monthrange
from datetime import date, datetime
from typing import Optional

from models import Company, Customer, Invoice, InvoiceItem, db
from utils.helpers import get_financial_year, parse_date, safe_float, safe_int

logger = logging.getLogger(__name__)


# ─── Shared openpyxl style helpers ────────────────────────────────────────────

def _require_openpyxl():
    """Import openpyxl or raise a clear error."""
    try:
        import openpyxl  # type: ignore
        return openpyxl
    except ImportError as exc:
        raise ImportError(
            "openpyxl is not installed. Run: pip install openpyxl"
        ) from exc


def _header_styles():
    """Return a dict of re-usable openpyxl style objects."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # type: ignore
    thin = Side(style="thin")
    return {
        "hdr_fill"   : PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid"),
        "hdr_font"   : Font(color="FFFFFF", bold=True, size=10),
        "total_fill" : PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
        "total_font" : Font(bold=True, size=10),
        "empty_fill" : PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid"),
        "border"     : Border(left=thin, right=thin, top=thin, bottom=thin),
        "center"     : Alignment(horizontal="center", vertical="center"),
        "right"      : Alignment(horizontal="right"),
    }


# ─── Excel import ─────────────────────────────────────────────────────────────

def import_invoices(filepath: str) -> dict:
    """Parse an Excel file and upsert Invoice records.

    Column layout (row 1 = header, row 2+ = data):

    ====  ==================  =========
    Col   Field               Required?
    ====  ==================  =========
    A     Invoice No (int)    Yes
    B     Date                Yes
    C     Customer Name       Yes
    D     Customer GSTIN      No
    E     Customer State      No
    F     Place of Supply     No
    G     Item Description    Yes
    H     Qty                 Yes
    I     Rate (₹)            Yes
    J     Unit                No (NOS)
    K     GST Rate (%)        No (18)
    ====  ==================  =========

    Rows sharing the same Invoice No are merged as line items on one invoice.
    An existing invoice with the same number + financial year is **updated**;
    otherwise a new invoice is **created**.

    Args:
        filepath: Absolute path to the ``.xlsx`` / ``.xls`` file.

    Returns:
        ``{"created": int, "updated": int, "skipped": int}``
    """
    openpyxl = _require_openpyxl()

    company = Company.query.first()
    company_prefix = company.gstin[:2] if company and company.gstin else "34"

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
    except Exception as exc:
        raise ValueError(f"Cannot open Excel file: {exc}") from exc

    # ── Group rows by invoice number ─────────────────────────────────────────
    invoice_rows: dict[int, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        try:
            inv_num = int(float(str(row[0]).strip()))
        except (ValueError, TypeError):
            continue

        if inv_num not in invoice_rows:
            invoice_rows[inv_num] = {
                "date"          : parse_date(row[1]),
                "customer_name" : str(row[2] or "").strip(),
                "customer_gstin": str(row[3] or "").strip(),
                "customer_state": str(row[4] or "").strip(),
                "place_of_supply": str(row[5] or "").strip(),
                "items"         : [],
            }

        desc = str(row[6] or "").strip()
        if desc:
            invoice_rows[inv_num]["items"].append({
                "description": desc,
                "qty"        : safe_int(row[7],    1),
                "rate"       : safe_float(row[8],  0.0),
                "unit"       : str(row[9]  or "NOS").strip() or "NOS",
                "gst_rate"   : safe_float(row[10], 18.0),
            })

    created = updated = skipped = 0

    for inv_num, data in invoice_rows.items():
        if not data["customer_name"]:
            logger.warning("Skipping invoice %s — no customer name.", inv_num)
            skipped += 1
            continue

        # Find or create customer
        customer = Customer.query.filter(
            Customer.name.ilike(data["customer_name"])
        ).first()
        if not customer:
            customer = Customer(
                name    = data["customer_name"],
                gstin   = data["customer_gstin"],
                state   = data["customer_state"],
                address = data["customer_state"],
            )
            db.session.add(customer)
            db.session.flush()

        inv_date = data["date"]
        fy       = get_financial_year(inv_date)
        fmt_num  = f"{fy}/{str(inv_num).zfill(3)}"
        is_intra = bool(
            data["customer_gstin"]
            and data["customer_gstin"][:2] == company_prefix
        )

        # Build line items + totals
        item_rows = data["items"] or [
            {"description": "General Supply", "qty": 1, "rate": 0.0, "unit": "NOS", "gst_rate": 18.0}
        ]
        total_basic = total_cgst = total_sgst = total_igst = 0.0
        inv_items: list[InvoiceItem] = []

        for it in item_rows:
            basic = round(it["qty"] * it["rate"], 2)
            gst   = round(basic * it["gst_rate"] / 100, 2)
            cgst  = round(gst / 2, 2) if is_intra else 0.0
            sgst  = round(gst / 2, 2) if is_intra else 0.0
            igst  = gst if not is_intra else 0.0
            total_basic += basic
            total_cgst  += cgst
            total_sgst  += sgst
            total_igst  += igst
            inv_items.append(InvoiceItem(
                description  = it["description"],
                qty          = it["qty"],
                rate         = it["rate"],
                unit         = it["unit"],
                gst_rate     = it["gst_rate"],
                basic_amount = basic,
                cgst_amount  = cgst,
                sgst_amount  = sgst,
                igst_amount  = igst,
                gst_amount   = gst,
                total_amount = basic + gst,
            ))

        total_gst   = total_cgst + total_sgst + total_igst
        grand_total = round(total_basic + total_gst, 2)

        existing = Invoice.query.filter_by(
            invoice_number_int=inv_num, financial_year=fy
        ).first()

        if existing:
            existing.date            = inv_date
            existing.customer_id     = customer.id
            existing.place_of_supply = data["place_of_supply"]
            existing.total_basic     = total_basic
            existing.total_cgst      = total_cgst
            existing.total_sgst      = total_sgst
            existing.total_igst      = total_igst
            existing.total_gst       = total_gst
            existing.grand_total     = grand_total
            existing.is_intra_state  = is_intra
            InvoiceItem.query.filter_by(invoice_id=existing.id).delete()
            for it in inv_items:
                it.invoice_id = existing.id
                db.session.add(it)
            updated += 1
        else:
            invoice = Invoice(
                invoice_number     = fmt_num,
                invoice_number_int = inv_num,
                date               = inv_date,
                financial_year     = fy,
                customer_id        = customer.id,
                place_of_supply    = data["place_of_supply"],
                total_basic        = total_basic,
                total_cgst         = total_cgst,
                total_sgst         = total_sgst,
                total_igst         = total_igst,
                total_gst          = total_gst,
                grand_total        = grand_total,
                percentage_cgst    = 9.0  if is_intra else 0.0,
                percentage_sgst    = 9.0  if is_intra else 0.0,
                percentage_igst    = 0.0  if is_intra else 18.0,
                is_intra_state     = is_intra,
            )
            db.session.add(invoice)
            db.session.flush()
            for it in inv_items:
                it.invoice_id = invoice.id
                db.session.add(it)
            created += 1

    db.session.commit()
    logger.info("Excel import: %d created, %d updated, %d skipped.", created, updated, skipped)
    return {"created": created, "updated": updated, "skipped": skipped}


# ─── GST report workbook ──────────────────────────────────────────────────────

def build_gst_report_workbook(month: int, year: int):
    """Build and return an ``openpyxl.Workbook`` for the GST monthly report.

    Fills invoice rows in sequential bill-number order.  Missing numbers
    in the range appear as greyed-out placeholder rows.

    Args:
        month: 1–12
        year:  e.g. 2025

    Returns:
        An ``openpyxl.Workbook`` ready to be streamed to the client.
    """
    openpyxl = _require_openpyxl()
    s = _header_styles()

    from_date = date(year, month, 1)
    to_date   = date(year, month, monthrange(year, month)[1])

    invoices = (
        Invoice.query
        .filter(Invoice.date >= from_date, Invoice.date <= to_date)
        .order_by(Invoice.invoice_number_int)
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = datetime(year, month, 1).strftime("GST %b %Y")

    headers = [
        "Bill No.", "Date", "Customer Name", "GSTIN", "Place of Supply",
        "Taxable Value", "CGST %", "CGST Amt", "SGST %", "SGST Amt",
        "IGST %", "IGST Amt", "Total Value",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill      = s["hdr_fill"]
        cell.font      = s["hdr_font"]
        cell.alignment = s["center"]
        cell.border    = s["border"]

    row_idx = 2
    if invoices:
        min_num  = invoices[0].invoice_number_int
        max_num  = invoices[-1].invoice_number_int
        inv_map  = {inv.invoice_number_int: inv for inv in invoices}
        fy       = invoices[0].financial_year

        for num in range(min_num, max_num + 1):
            if num in inv_map:
                inv = inv_map[num]
                row_data = [
                    inv.invoice_number,
                    inv.date.strftime("%d-%m-%Y"),
                    inv.customer.name,
                    inv.customer.gstin or "",
                    inv.place_of_supply or "",
                    round(inv.total_basic, 2),
                    inv.percentage_cgst if inv.is_intra_state else 0,
                    round(inv.total_cgst, 2),
                    inv.percentage_sgst if inv.is_intra_state else 0,
                    round(inv.total_sgst, 2),
                    inv.percentage_igst if not inv.is_intra_state else 0,
                    round(inv.total_igst, 2),
                    round(inv.grand_total, 2),
                ]
                for col, val in enumerate(row_data, 1):
                    cell           = ws.cell(row=row_idx, column=col, value=val)
                    cell.border    = s["border"]
                    if col in (6, 8, 10, 12, 13):
                        cell.alignment = s["right"]
            else:
                fmt = f"{fy}/{str(num).zfill(3)}"
                for col in range(1, 14):
                    cell        = ws.cell(row=row_idx, column=col,
                                          value=(fmt if col == 1 else ""))
                    cell.fill   = s["empty_fill"]
                    cell.border = s["border"]

            row_idx += 1

    # Totals row
    total_row = [
        "TOTAL", "", "", "", "",
        round(sum(i.total_basic  for i in invoices), 2),
        "",
        round(sum(i.total_cgst   for i in invoices), 2),
        "",
        round(sum(i.total_sgst   for i in invoices), 2),
        "",
        round(sum(i.total_igst   for i in invoices), 2),
        round(sum(i.grand_total  for i in invoices), 2),
    ]
    for col, val in enumerate(total_row, 1):
        cell        = ws.cell(row=row_idx, column=col, value=val)
        cell.fill   = s["total_fill"]
        cell.font   = s["total_font"]
        cell.border = s["border"]

    # Column widths
    for i, w in enumerate([14, 12, 28, 18, 16, 14, 8, 12, 8, 12, 8, 12, 14], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    return wb


# ─── Upload template workbook ─────────────────────────────────────────────────

def build_upload_template_workbook():
    """Return a pre-formatted ``openpyxl.Workbook`` for the invoice upload template.

    The workbook contains a header row and two sample invoices (one
    with two items, one with one item) so the user can see the expected format.

    Returns:
        An ``openpyxl.Workbook`` ready to be streamed to the client.
    """
    openpyxl = _require_openpyxl()
    from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice Template"

    headers = [
        "Invoice No*", "Date* (YYYY-MM-DD)", "Customer Name*",
        "Customer GSTIN", "Customer State", "Place of Supply",
        "Item Description*", "Qty*", "Rate*", "Unit", "GST Rate (%)",
    ]
    hdr_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=col, value=h)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal="center")

    today = datetime.today().strftime("%Y-%m-%d")
    sample = [
        [1, today, "ABC Pvt Ltd",  "33ABCDE1234F1Z5", "Tamil Nadu",  "Tamil Nadu",  "Item A",    2, 5000,  "NOS", 18],
        [1, today, "ABC Pvt Ltd",  "33ABCDE1234F1Z5", "Tamil Nadu",  "Tamil Nadu",  "Item B",    1, 3000,  "NOS", 12],
        [2, today, "XYZ Ltd",      "34XYZAB1234F1Z1", "Puducherry",  "Puducherry",  "Service X", 1, 10000, "NOS", 18],
    ]
    for row in sample:
        ws.append(row)

    for i, w in enumerate([12, 20, 25, 20, 15, 15, 25, 8, 10, 8, 12], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    return wb
